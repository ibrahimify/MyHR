"""
Import Data Page
- Upload CSV or Excel employee master data
- Validate rows before database writes
- Import valid rows only
- Provide a clean template for first-time company onboarding
"""

from datetime import datetime, timedelta
import csv
import os
import re
import zipfile
from xml.etree import ElementTree as ET

import qtawesome as qta
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QMessageBox, QSizePolicy
)

from src.core.i18n import t
from src.database.connection import (
    get_session, generate_employee_id, generate_commendation_ref,
    generate_sanction_ref, log_action, degree_to_title_name
)
from src.database.models import (
    Employee, Title, OrgUnit, Commendation,
    CommendationEmployee, Sanction
)


REQUIRED_COLUMNS = [
    "first_name", "last_name", "department", "degree",
    "position", "join_date", "base_salary"
]

OPTIONAL_COLUMNS = [
    "division", "unit", "team", "work_email", "work_phone",
    "personal_email", "phone", "address", "status",
    "manager_work_email", "commendation_months",
    "active_sanction_months", "sanction_type"
]

TEMPLATE_HEADERS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

DEGREES = {"BSc", "MSc", "PhD", "Other"}
STATUSES = {"active", "inactive", "on_leave"}
SANCTION_TYPES = {
    "verbal_warning": "verbal_warning",
    "verbal warning": "verbal_warning",
    "written_warning": "written_warning",
    "written warning": "written_warning",
    "suspension": "suspension",
    "final_warning": "final_warning",
    "final warning": "final_warning",
}

COLUMN_ALIASES = {
    "first": "first_name",
    "firstname": "first_name",
    "first_name": "first_name",
    "first_name_": "first_name",
    "first_name_required": "first_name",
    "last": "last_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "surname": "last_name",
    "email": "work_email",
    "work_email": "work_email",
    "work_email_address": "work_email",
    "company_email": "work_email",
    "department": "department",
    "dept": "department",
    "division": "division",
    "unit": "unit",
    "team": "team",
    "degree": "degree",
    "education": "degree",
    "position": "position",
    "job_title": "position",
    "title": "position",
    "join_date": "join_date",
    "joining_date": "join_date",
    "hire_date": "join_date",
    "start_date": "join_date",
    "salary": "base_salary",
    "base_salary": "base_salary",
    "monthly_salary": "base_salary",
    "work_phone": "work_phone",
    "phone": "phone",
    "mobile": "phone",
    "personal_email": "personal_email",
    "address": "address",
    "status": "status",
    "manager_email": "manager_work_email",
    "manager_work_email": "manager_work_email",
    "reports_to_email": "manager_work_email",
    "commendation_months": "commendation_months",
    "commendation_credit_months": "commendation_months",
    "active_sanction_months": "active_sanction_months",
    "sanction_months": "active_sanction_months",
    "sanction_type": "sanction_type",
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

PAGE_BG = "#f9fafb"
TEXT = "#030213"
MUTED = "#4b5563"
BORDER = "#e5e7eb"

CARD_SS = """
QFrame#Card {
    background: white;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
}
QFrame#Card QLabel {
    background: transparent;
    border: none;
}
"""

TABLE_SS = """
QTableWidget {
    background: white;
    alternate-background-color: white;
    border: none;
    gridline-color: #f3f4f6;
    font-size: 14px;
    color: #111827;
    outline: none;
    selection-background-color: #eff6ff;
}
QTableWidget::item {
    background: white;
    padding: 0 12px;
    border: none;
    border-bottom: 1px solid #f3f4f6;
    color: #111827;
}
QTableWidget::item:hover { background: #f9fafb; }
QTableWidget::item:selected { background: #eff6ff; color: #111827; }
QHeaderView::section {
    background: white;
    border: none;
    border-bottom: 1px solid #e5e7eb;
    padding: 0 12px;
    font-size: 13px;
    font-weight: 800;
    color: #030213;
    min-height: 50px;
    text-align: left;
}
QToolTip {
    background-color: #111827;
    color: white;
    border: 1px solid #374151;
    border-radius: 4px;
    padding: 6px 8px;
}
"""

MESSAGE_BOX_SS = """
QMessageBox { background: white; color: #111827; }
QMessageBox QLabel { color: #111827; background: transparent; font-size: 13px; }
QPushButton {
    background: white;
    color: #111827;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    min-width: 84px;
    min-height: 30px;
    font-weight: 600;
}
QPushButton:hover { background: #f3f4f6; }
QPushButton:default { background: #030213; color: white; border: none; }
"""


class ImportDataPage(QWidget):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.preview_data = []
        self.current_file = None
        self.step_widgets = []
        self.setStyleSheet(f"background: {PAGE_BG};")
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("border: none; background: transparent;")

        content = QWidget()
        content.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(content)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(0)

        title = QLabel("Import Employee Data")
        title.setStyleSheet(f"font-size: 30px; font-weight: 800; color: {TEXT}; background: transparent;")
        subtitle = QLabel("Bulk import employee records from CSV or Excel files")
        subtitle.setStyleSheet(f"font-size: 16px; color: {MUTED}; background: transparent;")
        layout.addWidget(title)
        layout.addSpacing(6)
        layout.addWidget(subtitle)
        layout.addSpacing(40)

        layout.addWidget(self._build_stepper())
        layout.addSpacing(30)

        body = QHBoxLayout()
        body.setSpacing(24)
        body.setAlignment(Qt.AlignTop)

        left = QVBoxLayout()
        left.setSpacing(20)
        left.setAlignment(Qt.AlignTop)
        left.addWidget(self._build_upload_card())
        left.addWidget(self._build_review_card())
        body.addLayout(left, 4)

        right = QVBoxLayout()
        right.setSpacing(20)
        right.setAlignment(Qt.AlignTop)
        right.addWidget(self._build_required_card())
        right.addWidget(self._build_cleaning_card())
        body.addLayout(right, 2)

        layout.addLayout(body)
        layout.addStretch()

        scroll.setWidget(content)
        outer.addWidget(scroll)
        self._set_step(0)

    def _build_stepper(self):
        frame = QFrame()
        frame.setStyleSheet("background: transparent; border: none;")
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        self.step_widgets = []
        for index, label in enumerate(["Upload File", "Validate Data", "Complete"]):
            item = QHBoxLayout()
            item.setSpacing(10)

            circle = QLabel(str(index + 1))
            circle.setFixedSize(40, 40)
            circle.setAlignment(Qt.AlignCenter)

            text = QLabel(label)
            text.setStyleSheet("font-size: 15px; font-weight: 800; color: #9ca3af; background: transparent;")

            item.addWidget(circle)
            item.addWidget(text)
            layout.addLayout(item)
            self.step_widgets.append((circle, text))

            if index < 2:
                line = QFrame()
                line.setFixedHeight(1)
                line.setStyleSheet("background: #d1d5db; border: none;")
                layout.addWidget(line, 1)

        return frame

    def _build_upload_card(self):
        self.upload_card = QFrame()
        self.upload_card.setObjectName("Card")
        self.upload_card.setMinimumHeight(430)
        self.upload_card.setStyleSheet(CARD_SS)
        layout = QVBoxLayout(self.upload_card)
        layout.setContentsMargins(48, 28, 48, 28)
        layout.setSpacing(12)
        layout.setAlignment(Qt.AlignCenter)

        icon_wrap = QLabel()
        icon_wrap.setFixedSize(76, 76)
        icon_wrap.setAlignment(Qt.AlignCenter)
        icon_wrap.setStyleSheet("background: #dbeafe; border-radius: 38px;")
        icon_wrap.setPixmap(qta.icon("fa5s.upload", color="#2563eb").pixmap(32, 32))
        layout.addWidget(icon_wrap, alignment=Qt.AlignCenter)

        upload_title = QLabel("Upload Employee Data File")
        upload_title.setAlignment(Qt.AlignCenter)
        upload_title.setStyleSheet(f"font-size: 22px; font-weight: 800; color: {TEXT}; background: transparent;")
        layout.addWidget(upload_title)

        upload_sub = QLabel("Supported formats: CSV, XLSX")
        upload_sub.setAlignment(Qt.AlignCenter)
        upload_sub.setStyleSheet(f"font-size: 16px; color: {MUTED}; background: transparent;")
        layout.addWidget(upload_sub)

        choose_btn = QPushButton("  Choose File")
        choose_btn.setIcon(qta.icon("fa5s.upload", color="white"))
        choose_btn.setIconSize(QSize(15, 15))
        choose_btn.setCursor(Qt.PointingHandCursor)
        choose_btn.setFixedSize(160, 50)
        choose_btn.setStyleSheet(_primary_button_ss())
        choose_btn.clicked.connect(self._choose_file)
        layout.addWidget(choose_btn, alignment=Qt.AlignCenter)

        self.file_label = QLabel("")
        self.file_label.setAlignment(Qt.AlignCenter)
        self.file_label.setStyleSheet("font-size: 13px; color: #2563eb; font-weight: 700; background: transparent;")
        layout.addWidget(self.file_label)

        line = QFrame()
        line.setFixedHeight(1)
        line.setStyleSheet("background: #e5e7eb; border: none;")
        layout.addSpacing(12)
        layout.addWidget(line)
        layout.addSpacing(14)

        template_btn = QPushButton("  Download Template File")
        template_btn.setIcon(qta.icon("fa5s.download", color="#111827"))
        template_btn.setIconSize(QSize(14, 14))
        template_btn.setCursor(Qt.PointingHandCursor)
        template_btn.setFixedSize(250, 44)
        template_btn.setStyleSheet(_secondary_button_ss())
        template_btn.clicked.connect(self._download_template)
        layout.addWidget(template_btn, alignment=Qt.AlignCenter)

        hint = QLabel("Use our template to ensure proper formatting")
        hint.setAlignment(Qt.AlignCenter)
        hint.setStyleSheet("font-size: 14px; color: #64748b; background: transparent;")
        layout.addWidget(hint)
        return self.upload_card

    def _build_review_card(self):
        self.review_card = QFrame()
        self.review_card.setObjectName("Card")
        self.review_card.setStyleSheet(CARD_SS)
        self.review_card.setVisible(False)
        layout = QVBoxLayout(self.review_card)
        layout.setContentsMargins(30, 28, 30, 30)
        layout.setSpacing(20)

        header = QHBoxLayout()
        icon = QLabel()
        icon.setPixmap(qta.icon("fa5s.clipboard-list", color="#2563eb").pixmap(18, 18))
        title = QLabel("Validation Results")
        title.setStyleSheet(f"font-size: 20px; font-weight: 800; color: {TEXT}; background: transparent;")
        header.addWidget(icon)
        header.addWidget(title)
        header.addStretch()
        layout.addLayout(header)

        self.stats_row = QHBoxLayout()
        self.stats_row.setSpacing(16)
        layout.addLayout(self.stats_row)

        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(9)
        self.preview_table.setHorizontalHeaderLabels([
            "Row", "Employee", "Department", "Position", "Degree",
            "Join Date", "Salary", "History", "Status"
        ])
        self.preview_table.setStyleSheet(TABLE_SS)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.preview_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.preview_table.setColumnWidth(0, 64)
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectRows)
        for col in range(self.preview_table.columnCount()):
            header_item = self.preview_table.horizontalHeaderItem(col)
            if header_item:
                header_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        layout.addWidget(self.preview_table)

        action_row = QHBoxLayout()
        action_row.addStretch()
        self.import_btn = QPushButton("  Import Valid Rows")
        self.import_btn.setIcon(qta.icon("fa5s.check-circle", color="white"))
        self.import_btn.setIconSize(QSize(15, 15))
        self.import_btn.setCursor(Qt.PointingHandCursor)
        self.import_btn.setFixedSize(230, 50)
        self.import_btn.setStyleSheet(_primary_button_ss())
        self.import_btn.clicked.connect(self._import)
        action_row.addWidget(self.import_btn)
        layout.addLayout(action_row)
        return self.review_card

    def _build_required_card(self):
        card = QFrame()
        card.setStyleSheet(
            "QFrame { background: #eff6ff; border-radius: 8px; border: 1px solid #bfdbfe; } "
            "QLabel { background: transparent; border: none; }"
        )
        layout = QVBoxLayout(card)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(9)

        head = QHBoxLayout()
        icon = QLabel()
        icon.setPixmap(qta.icon("fa5s.info-circle", color="#2563eb").pixmap(18, 18))
        title = QLabel("Required Columns")
        title.setStyleSheet("font-size: 17px; font-weight: 800; color: #1e40af; background: transparent;")
        head.addWidget(icon)
        head.addWidget(title)
        head.addStretch()
        layout.addLayout(head)

        for text in [
            "First Name",
            "Last Name",
            "Department",
            "Degree (BSc, MSc, PhD, Other)",
            "Position",
            "Join Date",
            "Base Salary",
        ]:
            layout.addWidget(_note_line(text, "#1d4ed8"))
        return card

    def _build_cleaning_card(self):
        card = QFrame()
        card.setStyleSheet(
            "QFrame { background: #fefce8; border-radius: 8px; border: 1px solid #fde047; } "
            "QLabel { background: transparent; border: none; }"
        )
        layout = QVBoxLayout(card)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(9)

        head = QHBoxLayout()
        icon = QLabel()
        icon.setPixmap(qta.icon("fa5s.exclamation-triangle", color="#d97706").pixmap(18, 18))
        title = QLabel("Data Cleaning Guide")
        title.setStyleSheet("font-size: 17px; font-weight: 800; color: #854d0e; background: transparent;")
        head.addWidget(icon)
        head.addWidget(title)
        head.addStretch()
        layout.addLayout(head)

        for text in [
            "Remove duplicate rows before import",
            "Validate work emails when provided",
            "Use standard department and team names",
            "Dates must use YYYY-MM-DD format",
            "Commendation months: 0, 1, 3, or 6",
            "Sanction months: 1-12",
        ]:
            layout.addWidget(_note_line(text, "#92400e"))
        return card

    def _set_step(self, step):
        for index, (circle, text) in enumerate(self.step_widgets):
            if index == step:
                circle.setStyleSheet(
                    "background: #2563eb; color: white; border-radius: 20px; "
                    "font-size: 16px; font-weight: 800;"
                )
                text.setStyleSheet("font-size: 15px; font-weight: 800; color: #2563eb; background: transparent;")
            elif index < step:
                circle.setStyleSheet(
                    "background: #dcfce7; color: #16a34a; border-radius: 20px; "
                    "font-size: 16px; font-weight: 800;"
                )
                text.setStyleSheet("font-size: 15px; font-weight: 800; color: #16a34a; background: transparent;")
            else:
                circle.setStyleSheet(
                    "background: #e5e7eb; color: #9ca3af; border-radius: 20px; "
                    "font-size: 16px; font-weight: 800;"
                )
                text.setStyleSheet("font-size: 15px; font-weight: 800; color: #9ca3af; background: transparent;")

    def _choose_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Employee Data File",
            "",
            "Employee Data Files (*.csv *.xlsx);;CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if not path:
            return

        self.current_file = path
        self.file_label.setText(os.path.basename(path))
        self._set_step(1)
        self._validate_file(path)

    def _validate_file(self, path):
        try:
            headers, raw_rows = self._read_file(path)
            if not headers:
                _critical(self, "Invalid File", "The selected file has no header row.")
                self._set_step(0)
                return

            canonical_headers = {_normalize_header(header) for header in headers}
            missing = [col for col in REQUIRED_COLUMNS if col not in canonical_headers]
            if missing:
                _critical(
                    self,
                    "Invalid File",
                    "Missing required columns: " + ", ".join(_title_column(col) for col in missing)
                )
                self._set_step(0)
                return

            existing_emails = self._existing_work_emails()
            seen_emails = set()
            rows = []

            for row_number, raw in enumerate(raw_rows, start=2):
                cleaned = _canonical_row(raw)
                if not any(str(value).strip() for value in cleaned.values()):
                    continue

                issues = self._validate_row(cleaned, row_number, seen_emails, existing_emails)
                normalized = self._normalized_row(cleaned, row_number, issues)
                rows.append(normalized)

            if not rows:
                _warning(self, "No Data", "The file contains headers but no employee rows.")
                self._set_step(0)
                return

            self.preview_data = rows
            self._show_preview(rows)
        except Exception as exc:
            _critical(self, t("error"), str(exc))
            self._set_step(0)

    def _read_file(self, path):
        extension = os.path.splitext(path)[1].lower()
        if extension == ".csv":
            return self._read_csv(path)
        if extension == ".xlsx":
            return self._read_xlsx(path)
        raise ValueError("Supported file formats are CSV and XLSX.")

    def _read_csv(self, path):
        with open(path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            headers = reader.fieldnames or []
            return headers, list(reader)

    def _read_xlsx(self, path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            return self._read_xlsx_fallback(path)

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return [], []
            headers = [_cell_to_text(value) for value in rows[0]]
            records = []
            for values in rows[1:]:
                if not any(_cell_to_text(value) for value in values):
                    continue
                records.append({
                    headers[index]: _cell_to_text(value)
                    for index, value in enumerate(values)
                    if index < len(headers) and headers[index]
                })
            return headers, records
        finally:
            workbook.close()

    def _read_xlsx_fallback(self, path):
        namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with zipfile.ZipFile(path) as archive:
            shared_strings = []
            if "xl/sharedStrings.xml" in archive.namelist():
                shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for item in shared_root.findall("a:si", namespace):
                    shared_strings.append("".join(
                        text_node.text or "" for text_node in item.findall(".//a:t", namespace)
                    ))

            sheet_name = "xl/worksheets/sheet1.xml"
            if sheet_name not in archive.namelist():
                worksheets = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet"))
                if not worksheets:
                    return [], []
                sheet_name = worksheets[0]

            sheet_root = ET.fromstring(archive.read(sheet_name))
            rows = []
            for row_node in sheet_root.findall(".//a:sheetData/a:row", namespace):
                values = []
                for cell in row_node.findall("a:c", namespace):
                    ref = cell.attrib.get("r", "")
                    col_index = _excel_col_index(ref)
                    while len(values) < col_index:
                        values.append("")
                    values[col_index - 1] = _xlsx_cell_value(cell, shared_strings, namespace)
                rows.append(values)

        if not rows:
            return [], []

        headers = [_cell_to_text(value) for value in rows[0]]
        records = []
        for row in rows[1:]:
            if not any(_cell_to_text(value) for value in row):
                continue
            records.append({
                headers[index]: _cell_to_text(value)
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            })
        return headers, records

    def _existing_work_emails(self):
        session = get_session()
        try:
            values = session.query(Employee.work_email).filter(Employee.work_email.isnot(None)).all()
            return {email.lower() for (email,) in values if email}
        finally:
            session.close()

    def _validate_row(self, row, row_number, seen_emails, existing_emails):
        issues = []
        for column in REQUIRED_COLUMNS:
            if not _value(row.get(column)):
                issues.append(f"Missing {_title_column(column)}")

        degree = _normalize_degree(row.get("degree"))
        if _value(row.get("degree")) and not degree:
            issues.append("Degree must be BSc, MSc, PhD, or Other")

        join_date = _parse_date(row.get("join_date"))
        if _value(row.get("join_date")) and not join_date:
            issues.append("Join Date must be YYYY-MM-DD")
        elif join_date and join_date > datetime.utcnow():
            issues.append("Join Date cannot be in the future")

        salary = _parse_float(row.get("base_salary"))
        if _value(row.get("base_salary")) and salary is None:
            issues.append("Base Salary must be a number")
        elif salary is not None and salary <= 0:
            issues.append("Base Salary must be greater than zero")

        status = _value(row.get("status")) or "active"
        if status and status not in STATUSES:
            issues.append("Status must be active, inactive, or on_leave")

        work_email = _value(row.get("work_email")).lower()
        if work_email:
            if not EMAIL_RE.match(work_email):
                issues.append("Work Email format is invalid")
            elif work_email in existing_emails:
                issues.append("Work Email already exists in MyHR")
            elif work_email in seen_emails:
                issues.append("Work Email is duplicated in this file")
            else:
                seen_emails.add(work_email)

        manager_email = _value(row.get("manager_work_email")).lower()
        if manager_email and not EMAIL_RE.match(manager_email):
            issues.append("Manager Work Email format is invalid")

        comm_months = _parse_int(row.get("commendation_months"), default=0)
        if comm_months not in {0, 1, 3, 6}:
            issues.append("Commendation Months must be 0, 1, 3, or 6")

        sanction_months = _parse_int(row.get("active_sanction_months"), default=0)
        if sanction_months < 0 or sanction_months > 12:
            issues.append("Active Sanction Months must be between 0 and 12")

        sanction_type = _value(row.get("sanction_type")).lower()
        if sanction_months > 0 and sanction_type and sanction_type not in SANCTION_TYPES:
            issues.append("Sanction Type must be Verbal Warning, Written Warning, Suspension, or Final Warning")

        return issues

    def _normalized_row(self, row, row_number, issues):
        degree = _normalize_degree(row.get("degree")) or _value(row.get("degree"))
        status = _value(row.get("status")) or "active"
        comm_months = _parse_int(row.get("commendation_months"), default=0)
        sanction_months = _parse_int(row.get("active_sanction_months"), default=0)
        return {
            "row": row_number,
            "first_name": _value(row.get("first_name")),
            "last_name": _value(row.get("last_name")),
            "department": _value(row.get("department")),
            "division": _value(row.get("division")),
            "unit": _value(row.get("unit")),
            "team": _value(row.get("team")),
            "degree": degree,
            "position": _value(row.get("position")),
            "join_date": _value(row.get("join_date")),
            "base_salary": _value(row.get("base_salary")),
            "work_email": _value(row.get("work_email")),
            "work_phone": _value(row.get("work_phone")),
            "phone": _value(row.get("phone")),
            "personal_email": _value(row.get("personal_email")),
            "address": _value(row.get("address")),
            "status_value": status,
            "manager_work_email": _value(row.get("manager_work_email")),
            "commendation_months": comm_months,
            "active_sanction_months": sanction_months,
            "sanction_type": SANCTION_TYPES.get(_value(row.get("sanction_type")).lower(), "written_warning"),
            "status": "error" if issues else "valid",
            "issues": issues,
        }

    def _show_preview(self, rows):
        valid = [row for row in rows if row["status"] == "valid"]
        errors = [row for row in rows if row["status"] == "error"]
        self._set_step(1)
        self.review_card.setVisible(True)
        self._populate_stats(rows, valid, errors)
        self._populate_preview(rows)
        self.import_btn.setEnabled(bool(valid))
        self.import_btn.setText(f"  Import {len(valid)} Valid Row{'s' if len(valid) != 1 else ''}")
        self.review_card.adjustSize()

    def _populate_stats(self, rows, valid, errors):
        while self.stats_row.count():
            item = self.stats_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for label, value, color, bg, icon_name in [
            ("Total Rows", len(rows), "#2563eb", "#dbeafe", "fa5s.file-alt"),
            ("Valid Rows", len(valid), "#16a34a", "#dcfce7", "fa5s.check-circle"),
            ("Needs Fix", len(errors), "#dc2626", "#fee2e2", "fa5s.exclamation-triangle"),
        ]:
            self.stats_row.addWidget(_stat_card(label, value, color, bg, icon_name))
        self.stats_row.addStretch()

    def _populate_preview(self, rows):
        self.preview_table.setRowCount(len(rows))
        self.preview_table.setMinimumHeight(112 + (52 * min(max(len(rows), 1), 8)))

        for index, row in enumerate(rows):
            self.preview_table.setRowHeight(index, 52)
            employee = f"{row['first_name']} {row['last_name']}".strip()
            history = []
            if row["commendation_months"]:
                history.append(f"-{row['commendation_months']} mo commendation")
            if row["active_sanction_months"]:
                history.append(f"+{row['active_sanction_months']} mo sanction")
            history_text = ", ".join(history) if history else "-"

            values = [
                str(row["row"]),
                employee,
                row["department"],
                row["position"],
                row["degree"],
                row["join_date"],
                row["base_salary"],
                history_text,
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setToolTip(value)
                if col == 0:
                    item.setForeground(QColor("#6b7280"))
                self.preview_table.setItem(index, col, item)

            if row["status"] == "valid":
                status_text = "Valid"
                status = QTableWidgetItem(status_text)
                status.setIcon(qta.icon("fa5s.check-circle", color="#16a34a"))
                status.setForeground(QColor("#16a34a"))
                status.setToolTip(status_text)
            else:
                issue_text = "; ".join(row["issues"])
                status = QTableWidgetItem("Fix Required")
                status.setIcon(qta.icon("fa5s.exclamation-triangle", color="#dc2626"))
                status.setForeground(QColor("#dc2626"))
                status.setToolTip(issue_text)
                for col in range(8):
                    item = self.preview_table.item(index, col)
                    if item:
                        item.setBackground(QColor("#fef2f2"))
            self.preview_table.setItem(index, 8, status)

    def _import(self):
        valid_rows = [row for row in self.preview_data if row["status"] == "valid"]
        if not valid_rows:
            return

        confirm = _question(
            self,
            "Confirm Import",
            f"Import {len(valid_rows)} validated employee record(s)?"
        )
        if confirm != QMessageBox.Yes:
            return

        session = get_session()
        imported = 0
        errors = []
        created = []

        try:
            for row in valid_rows:
                try:
                    title_name = degree_to_title_name(row["degree"])
                    title = session.query(Title).filter_by(name=title_name).first()
                    if not title:
                        errors.append(f"Row {row['row']}: Title {title_name} was not found")
                        continue

                    work_email = row["work_email"].lower() if row["work_email"] else None
                    if work_email:
                        exists = session.query(Employee).filter(Employee.work_email == work_email).first()
                        if exists:
                            errors.append(f"Row {row['row']}: Work email already exists")
                            continue

                    join_date = _parse_date(row["join_date"])
                    employee = Employee(
                        employee_id=generate_employee_id(session),
                        first_name=row["first_name"],
                        last_name=row["last_name"],
                        degree=row["degree"],
                        position=row["position"],
                        join_date=join_date,
                        base_salary=float(row["base_salary"]),
                        work_email=work_email,
                        work_phone=row["work_phone"] or None,
                        phone=row["phone"] or None,
                        personal_email=row["personal_email"] or None,
                        address=row["address"] or None,
                        status=row["status_value"],
                        title_id=title.id,
                        org_unit_id=self._org_unit_id(session, row),
                    )
                    session.add(employee)
                    session.flush()
                    created.append((employee, row))
                    imported += 1
                except Exception as exc:
                    errors.append(f"Row {row['row']}: {exc}")

            self._wire_managers(session, created)
            self._import_optional_history(session, created)

            log_action(
                session=session,
                action="import.bulk_employees",
                performed_by_id=self.user.id,
                target_table="employee",
                target_id=None,
                description=f"Bulk employee import completed: {imported} employees imported"
            )
            session.commit()

            self._set_step(2)
            message = f"Successfully imported {imported} employee record(s)."
            if errors:
                message += "\n\nRows needing review:\n" + "\n".join(errors[:8])
            _information(self, t("import_success"), message)
            self._reset_page()
        except Exception as exc:
            session.rollback()
            _critical(self, t("error"), str(exc))
        finally:
            session.close()

    def _wire_managers(self, session, created):
        email_lookup = {}
        for employee, _ in created:
            if employee.work_email:
                email_lookup[employee.work_email.lower()] = employee

        existing = session.query(Employee).filter(Employee.work_email.isnot(None)).all()
        for employee in existing:
            if employee.work_email:
                email_lookup.setdefault(employee.work_email.lower(), employee)

        for employee, row in created:
            manager_email = row["manager_work_email"].lower() if row["manager_work_email"] else ""
            manager = email_lookup.get(manager_email)
            if manager and manager.id != employee.id:
                employee.reports_to_id = manager.id

    def _import_optional_history(self, session, created):
        for employee, row in created:
            issued_at = _event_date_for(employee.join_date)

            if row["commendation_months"]:
                category = {1: 1, 3: 2, 6: 3}[row["commendation_months"]]
                comm = Commendation(
                    commendation_ref=generate_commendation_ref(session),
                    title="Imported Prior Commendation",
                    description="Imported from employee onboarding file.",
                    category=category,
                    months_impact=-row["commendation_months"],
                    is_team_award=False,
                    issued_by_id=self.user.id,
                    issued_at=issued_at,
                )
                session.add(comm)
                session.flush()
                session.add(CommendationEmployee(commendation_id=comm.id, employee_id=employee.id))

            if row["active_sanction_months"]:
                sanction = Sanction(
                    sanction_ref=generate_sanction_ref(session),
                    employee_id=employee.id,
                    sanction_type=row["sanction_type"],
                    reason="Imported active disciplinary record.",
                    delay_months=row["active_sanction_months"],
                    issued_by_id=self.user.id,
                    issued_at=issued_at,
                    is_resolved=False,
                )
                session.add(sanction)

    def _org_unit_id(self, session, row):
        department_name = row["department"] or "Unassigned"
        division_name = row["division"]
        unit_name = row["unit"]
        team_name = row["team"]

        parent = None
        if division_name:
            company = _get_or_create_org_unit(session, "Imported Company", "organization", None)
            parent = _get_or_create_org_unit(session, division_name, "division", company.id)

        department = _get_or_create_org_unit(
            session, department_name, "department", parent.id if parent else None
        )
        current = department
        if unit_name:
            current = _get_or_create_org_unit(session, unit_name, "unit", current.id)
        if team_name:
            current = _get_or_create_org_unit(session, team_name, "team", current.id)
        return current.id

    def _reset_page(self):
        self.preview_data = []
        self.current_file = None
        self.file_label.setText("")
        self.review_card.setVisible(False)
        self.preview_table.setRowCount(0)
        while self.stats_row.count():
            item = self.stats_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._set_step(0)

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Template",
            "employee_import_template.csv",
            "CSV Files (*.csv)"
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"

        try:
            with open(path, "w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(TEMPLATE_HEADERS)
                writer.writerow([
                    "James", "Wilson", "Engineering", "BSc", "Software Engineer I",
                    "2023-04-10", "3200", "Engineering", "API Unit", "Payments Team",
                    "james.wilson@company.com", "+36 20 110 2200",
                    "james.personal@example.com", "+36 30 110 2200", "Budapest",
                    "active", "lead.engineer@company.com", "1", "0", ""
                ])
                writer.writerow([
                    "Nora", "Szabo", "Human Resources", "MSc", "HR Business Partner",
                    "2022-09-01", "3600", "Corporate Services", "People Operations", "Payroll Team",
                    "nora.szabo@company.com", "+36 20 330 4400",
                    "", "", "Budapest", "active", "", "0", "0", ""
                ])
                writer.writerow([
                    "David", "Chen", "Quality Assurance", "BSc", "QA Engineer",
                    "2021-02-15", "3000", "Product Delivery", "Automation QA", "Regression Team",
                    "david.chen@company.com", "+36 20 550 6600",
                    "", "", "Szeged", "active", "qa.manager@company.com", "0", "2", "verbal_warning"
                ])
            _information(self, t("success"), f"Template saved to:\n{path}")
        except Exception as exc:
            _critical(self, t("error"), str(exc))


def _normalize_header(header):
    text = str(header or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return COLUMN_ALIASES.get(text, text)


def _canonical_row(raw):
    output = {}
    for key, value in (raw or {}).items():
        canonical = _normalize_header(key)
        if not canonical:
            continue
        text = _cell_to_text(value)
        if canonical not in output or text:
            output[canonical] = text
    return output


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _excel_col_index(ref):
    letters = "".join(ch for ch in ref if ch.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index, 1)


def _xlsx_cell_value(cell, shared_strings, namespace):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(
            text_node.text or "" for text_node in cell.findall(".//a:t", namespace)
        )

    value_node = cell.find("a:v", namespace)
    value = value_node.text if value_node is not None else ""
    if cell_type == "s" and value:
        try:
            return shared_strings[int(value)]
        except (ValueError, IndexError):
            return value
    return value or ""


def _value(value):
    return str(value or "").strip()


def _title_column(column):
    return column.replace("_", " ").title()


def _normalize_degree(value):
    text = _value(value).lower()
    mapping = {"bsc": "BSc", "msc": "MSc", "phd": "PhD", "other": "Other"}
    return mapping.get(text)


def _parse_date(value):
    text = _value(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return None


def _parse_float(value):
    text = _value(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value, default=0):
    text = _value(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return -999


def _event_date_for(join_date):
    now = datetime.utcnow()
    candidate = now - timedelta(days=45)
    if join_date and candidate < join_date:
        return min(now, join_date + timedelta(days=30))
    return candidate


def _get_or_create_org_unit(session, name, unit_type, parent_id):
    query = session.query(OrgUnit).filter(
        OrgUnit.name == name,
        OrgUnit.unit_type == unit_type
    )
    if parent_id is None:
        query = query.filter(OrgUnit.parent_id.is_(None))
    else:
        query = query.filter(OrgUnit.parent_id == parent_id)

    unit = query.first()
    if unit:
        return unit

    unit = OrgUnit(name=name, unit_type=unit_type, parent_id=parent_id)
    session.add(unit)
    session.flush()
    return unit


def _stat_card(label, value, color, bg, icon_name):
    card = QFrame()
    card.setObjectName("Card")
    card.setFixedHeight(86)
    card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
    card.setStyleSheet(CARD_SS)
    layout = QHBoxLayout(card)
    layout.setContentsMargins(20, 0, 20, 0)
    layout.setSpacing(12)

    icon = QLabel()
    icon.setFixedSize(44, 44)
    icon.setAlignment(Qt.AlignCenter)
    icon.setStyleSheet(f"background: {bg}; border-radius: 8px;")
    icon.setPixmap(qta.icon(icon_name, color=color).pixmap(20, 20))

    texts = QVBoxLayout()
    texts.setSpacing(0)
    texts.setAlignment(Qt.AlignVCenter)
    title = QLabel(label)
    title.setStyleSheet("font-size: 13px; color: #374151; background: transparent;")
    number = QLabel(str(value))
    number.setStyleSheet("font-size: 24px; font-weight: 800; color: #030213; background: transparent;")
    texts.addWidget(title)
    texts.addWidget(number)

    layout.addWidget(icon)
    layout.addLayout(texts)
    return card


def _note_line(text, color):
    label = QLabel("&bull; " + text)
    label.setTextFormat(Qt.RichText)
    label.setWordWrap(False)
    label.setStyleSheet(f"font-size: 13px; color: {color}; background: transparent;")
    return label


def _primary_button_ss():
    return """
QPushButton {
    background: #030213;
    color: white;
    border: none;
    border-radius: 8px;
    font-size: 14px;
    font-weight: 800;
}
QPushButton:hover { background: #111827; }
QPushButton:disabled { background: #d1d5db; color: #9ca3af; }
"""


def _secondary_button_ss():
    return """
QPushButton {
    background: white;
    color: #111827;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    font-size: 14px;
    font-weight: 700;
}
QPushButton:hover { background: #f3f4f6; }
"""


def _styled_message_box(parent, icon, title, text, buttons=QMessageBox.Ok, default_button=QMessageBox.Ok):
    box = QMessageBox(parent)
    box.setIcon(icon)
    box.setWindowTitle(title)
    box.setText(text)
    box.setStandardButtons(buttons)
    box.setDefaultButton(default_button)
    box.setStyleSheet(MESSAGE_BOX_SS)
    return box.exec()


def _warning(parent, title, text):
    return _styled_message_box(parent, QMessageBox.Warning, title, text)


def _critical(parent, title, text):
    return _styled_message_box(parent, QMessageBox.Critical, title, text)


def _information(parent, title, text):
    return _styled_message_box(parent, QMessageBox.Information, title, text)


def _question(parent, title, text):
    return _styled_message_box(
        parent, QMessageBox.Question, title, text,
        QMessageBox.Yes | QMessageBox.No, QMessageBox.No
    )
